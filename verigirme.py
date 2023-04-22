import openpyxl
dosya=openpyxl.load_workbook("./data.xlsx")
sayfa=dosya["notlar"]
sayfa.cell(row=1,column=5,value="Harf Notu")
dosya.save("./data.xlsx")
satir_sayfasi=sayfa.max_row
sutun_sayisi=sayfa.max_column
for i in range (2,satir_sayfasi+1):
    for j in range (4,sutun_sayisi):
        print(sayfa.cell(i,j).value)
        if sayfa.cell(i,j).value>=85:
            sayfa.cell(row=i,column=5,value="A")
        elif sayfa.cell(i,j).value>=70:
            sayfa.cell(row=i,column=5,value="B")
        elif sayfa.cell(i,j).value>=60:
            sayfa.cell(row=i,column=5,value="C")
        elif sayfa.cell(i,j).value>=50:
            sayfa.cell(row=i,column=5,value="D")
        else:
            sayfa.cell(row=i,column=5,value="F")
dosya.save("./data.xlsx")

