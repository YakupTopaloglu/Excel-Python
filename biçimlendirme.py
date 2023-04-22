import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.fills import PatternFill
dosya=openpyxl.load_workbook("./data.xlsx")
sayfa=dosya["notlar"]
sayfa.cell(row=1,column=5,value="Harf Notu")
sayfa.cell(row=1,column=5).font=Font(name="calibri",size=16)
dosya.save("./data.xlsx")
satir_sayfasi=sayfa.max_row
sutun_sayisi=sayfa.max_column
for i in range (2,satir_sayfasi+1):
    for j in range (4,sutun_sayisi):
        if sayfa.cell(i,j).value>=85:
            sayfa.cell(row=i,column=5,value="A")
            sayfa.cell(row=i,column=5).font=Font(name="calibri",size=16,color="3A4B54")
            sayfa.cell(row=i,column=5).fill=PatternFill("solid","0CFF00")
        elif sayfa.cell(i,j).value>=70:
            sayfa.cell(row=i,column=5,value="B")
            sayfa.cell(row=i,column=5).font=Font(name="calibri",size=16,color="3A4B54")
            sayfa.cell(row=i,column=5).fill=PatternFill("solid","00FFA6")
        elif sayfa.cell(i,j).value>=60:
            sayfa.cell(row=i,column=5,value="C")
            sayfa.cell(row=i,column=5).font=Font(name="calibri",size=16,color="3A4B54")
            sayfa.cell(row=i,column=5).fill=PatternFill("solid","D0D908")
        elif sayfa.cell(i,j).value>=50:
            sayfa.cell(row=i,column=5,value="D")
            sayfa.cell(row=i,column=5).font=Font(name="calibri",size=16,color="3A4B54")
            sayfa.cell(row=i,column=5).fill=PatternFill("solid","D97708")
        else:
            sayfa.cell(row=i,column=5,value="F")
            sayfa.cell(row=i,column=5).font=Font(name="calibri",size=16,color="3A4B54")
            sayfa.cell(row=i,column=5).fill=PatternFill("solid","FF0000")
dosya.save("./data.xlsx")

