import openpyxl

def liste(dosyayeri,sayfaadi):
    dosya=openpyxl.load_workbook(dosyayeri)
    sayfa=dosya[sayfaadi]
    satir_sayisi=sayfa.max_row
    sutun_sayisi=sayfa.max_column
    data=[]
    for i in range(2,satir_sayisi+1):
        satir=[]
        for j in range(1,sutun_sayisi+1):
            satir.append(sayfa.cell(i,j).value)
        data.append(satir)
    return data
        
deneme=liste("./data.xlsx","notlar")
print(deneme)