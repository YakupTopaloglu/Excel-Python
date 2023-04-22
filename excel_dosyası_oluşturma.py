import openpyxl
dosya=openpyxl.Workbook()
sayfa=dosya.active
sayfa.title="Hatalar"
sayfa["A1"].value="Python"
sayfa["A2"].value="Python"
dosya.create_sheet("ikinci",1)
dosya.create_sheet("üçüncü",2)
dosya.save("./yeni.xlsx")