#dosya-sayfa-satır/sutun-hucre-deger
import openpyxl
dosya=openpyxl.load_workbook("./Data.xlsx")
print(dosya)
print("Aktif sayfa: "+dosya.active.title)
print(dosya.sheetnames)
sayfa=dosya["Sayfa1"]
deger=sayfa["B4"].value
print(deger)
#*************
sayfalar=dosya["Sayfa1"]
veri=sayfalar.cell(4,2).value
print(veri)